import shutil
import openpyxl
import create_folders
import pandas as pd
import numpy as np
import datetime, os
import win32com.client as win32
from openpyxl.styles import PatternFill , Alignment , Font
import open_files

#df=pd.read_excel("C:\\Users\\atharva.kabra\\OneDrive - Infosys Limited\\Desktop\\report\\offboarding\\Master Access Management Tracker 30 June 2023.xlsx")
df = pd.read_excel(open_files.master_path , sheet_name= 'Sheet1')

df = df[df['Approval email\n(Yes/No)'].isin(['Yes','No'])]
df =df[df['Attrition Reason'].isin(['Role Ended','Resource quit the vendor organization',
                                    'Resource extended leave - medical/emergency',
                                    'Closed By FG Team',
                                    "Incorrect on-boarding",
                                    "Role ended",
                                    "Inadequate notification period",
                                    "Resource performance issues",
                                    "SP Initiated Resource performance issues",
                                    "Bank Initiated Resource fulfillment issue",
                                    "Resource performance issues",
                                    "Subcontractor Tenure Limitation",
                                    "Resource quit the vendor organization",
                                    "SP Initiated Resource performance issues",
                                    "Resource moved out of BAC account"])]


#calculating No of Days 
df[['Offboarding Notification Mail sent or received on','End Date']] = df[['Offboarding Notification Mail sent or received on','End Date']].apply(pd.to_datetime) #if conversion required
df['No. of Days'] = (df['End Date'] - df['Offboarding Notification Mail sent or received on']).dt.days
#status of employes 
df['Status'] = df.apply(lambda row: "Less than or equal to 1 Week" if row["No. of Days"]<=7
                                else ("Between 1-2 weeks" if row["No. of Days"]>=8 and row["No. of Days"]<15
                                        else ("Greater than 2 weeks" )), axis=1)

df = df.drop(columns=[
    'ITP Approval received', 'Training screen Shot ',
    'Is Movement across GDCE ID - ITP requested for uninterrupted access',
    'Job Level', 'Person ID',
    'Resource Mailcode', 'Any Pending Timesheets or Revisions',
    'Contractor end date',
    'Actual Reason', 'Redeployment\n(Yes/No)',
    'Resource available to redeploy (Yes/No) - Selected in FG/Template',
    'If no, Reason for unavailability for redeployment ',
    'If yes, SOW Name for new onboarding',
    'If yes, Work order # of new onboarding (BACGWOxxxxxx)',
    'Available in Redeployment portal', 'GT Level Report',
    'Does Onsite resource holding Bank asset\n(Yes/No)',
    'Date on when the request submitted',
    'Request number for device submission', 'Status on device submission',
    'FG End Date Updated on (Revision Submission Date)',
    'Physical access revoked on', 'Domain Access Revoked on',
    'FG Closed Date', 'Closed in FG\n(Yes/No)',
    'AHD request number for disabling Domain access', 'FG access',
    'Reason for short Notice (Less than 2 weeks)', 'Vendor Comments',
    'Replacement Resource Name', 'Replacement Resource EMP #', 'Worker ID',
    'On boarding date / registration date', 'QA (Audit) \nComment ',
    'Audited By', 'Reviewed on\n(Date)',
    'Reported in weekly report on\n(Date)',  'CIO Function',
    'OPS ACTUAL REASON', ])

# df.drop(columns=df.columns[(df == 'Salty').any()])
df.drop(df.loc[df['SOW ID'] == 'BACGTQ00000668'].index, inplace=True)
df.drop(df.loc[df['Rehire Status'] == 'No'].index, inplace=True)


df['End Date Month'] =df['End Date'].dt.strftime("%b")
# print(df['End Date Month'])   

#convert empty values to empty strings
df = df.replace({pd.NA: ''})

# print(df["""Reason with ITP's Approval\n(If Rehire status is "NO")"""].value_counts())
# print(df.columns)
# Worker Infosys Email	Offboarding Notification Mail sent or received on
# last_col = df.columns[-1]
# print(last_col)
# df.insert(14, last_col,df[last_col])
# df.drop(last_col,axis=1 , inplace=True)
neworder = ["Email ID of Requester","EMP#",	"Worker Infosys Email","Offboarding Notification Mail sent or received on","Off boarding Approval received on",	"Approval email\n(Yes/No)",	"Emp DU",	"Emp Unit","PM",	"DM",	"SEM",	"Location",	"Contractor Id",	"Contractor Name",	"SOW ID",	"End Date", "End Date Month",	"Attrition category",	"Attrition Reason",	"Rehire Status",	"""Reason with ITP's Approval\n(If Rehire status is "NO")""",  	"GDCE ID",	"Skill 2",	"Skill 3",	"No. of Days",	"Status"]
df=df.reindex(columns=neworder)
# print(df.columns)

#pivot table
grouped_data = df.groupby(['SEM','Status']).size().unstack(fill_value=0)
dff = grouped_data.assign(Total = grouped_data.sum(axis=1))
dff.loc['Grand Total'] = dff.sum(axis=0)
a =dff.iloc[-1].values
less = a[0]/a[-1]*100
bet = a[1]/a[-1]*100
greter =a[2]/a[-1]*100
print(dff)

import datetime
today = datetime.date.today()
first = today.replace(day=1)
last_month = first 
last_month.strftime("%b")
last_last_month = last_month - datetime.timedelta(days=31)
last_last_month.strftime("%b")
# dff = dff.to_html(index=False, na_rep ='').replace('<th>','<th style= "background-color:yellow">')
# dff.style.set_properties({'background-color': 'black','color': 'green'})

# dff.to_excel("C:\\Users\\atharva.kabra\\OneDrive - Infosys Limited\Desktop\\ofboarding\\attachment.xlsx",sheet_name="pivot")

os.chdir(create_folders.output_path)
with pd.ExcelWriter(f"Offboarding_Report,{datetime.datetime.now().date()}.xlsx") as writer:
    #dff.to_excel(writer, sheet_name="Pivot",)
    df.to_excel(writer, sheet_name="Output_Data", index=False)
    output_file_path = os.path.join(create_folders.output_path,f"Offboarding_Report,{datetime.datetime.now().date()}.xlsx")

os.chdir(create_folders.output_path)
with pd.ExcelWriter(f"Pivot_Report,{datetime.datetime.now().date()}.xlsx") as writer:
    # dff.to_html(index=False, na_rep ='').replace('<th>','<th style= "background-color:yellow">')
    dff.to_excel(writer, sheet_name="Pivot",)
    #df.to_excel(writer, sheet_name="Output_Data", index=False)
    Pivot_file_path = os.path.join(create_folders.output_path,f"Pivot_Report,{datetime.datetime.now().date()}.xlsx")



# ----------------------------- Colors & Alignments ------------------------------      
op_report = openpyxl.load_workbook(output_file_path)
op_report_sheet = op_report.worksheets[0]
# op_report_sheet.insert_rows(1)
# op_report_sheet.cell(1,3).value = "Offboarding Report"

#apply filter for row 2
op_report_sheet.auto_filter.ref = f"A1:{op_report_sheet.cell(row =2,column= op_report_sheet.max_column).coordinate}"

# add color to headers
for i in range(1,op_report_sheet.max_column+1):

    op_report_sheet.cell(row=1,column= i).fill = PatternFill(start_color = "99CCFF", end_color = "99CCFF", fill_type = "solid")

# Specify alignment options for rows and columns
alignment = Alignment(horizontal='center', vertical='center')

# Align rows
for row in op_report_sheet.iter_rows():
    for cell in row:
        cell.alignment = alignment

# Align columns
for column in op_report_sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
            op_report_sheet.column_dimensions[column_letter].width = max_length + 2
            
    for cell in column:
        cell.alignment = alignment

font = Font(size = 12)
# op_report_sheet.cell(1,3).font = font
# op_report_sheet.cell(row=1,column= 3).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
# op_report_sheet.merge_cells('C1:D1')

# save file (interim adn yellow sheets)
op_report.save(output_file_path)



#--------------------------- Move old files to backup folder ------------------------------------
for filename in os.listdir(create_folders.output_path):
    if not filename.startswith("~$") and  str(datetime.datetime.now().date()) not in filename: 
        shutil.move(os.path.join(create_folders.output_path, filename), create_folders.output_path_backup)